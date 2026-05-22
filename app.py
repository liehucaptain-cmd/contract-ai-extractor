"""
合同智能提取工具
两段式：Surya OCR → 原文展示 → Ollama文本模型提取字段 → Excel台账
"""

import sys, os, logging, traceback, json, base64, re, time
from io import BytesIO
from datetime import datetime
from pathlib import Path

# ===================== 清除系统代理（必须在网络库 import 之前） =====================
for _key in ["http_proxy", "https_proxy", "HTTP_PROXY", "HTTPS_PROXY", "all_proxy", "ALL_PROXY"]:
    os.environ.pop(_key, None)
os.environ.setdefault("no_proxy", "localhost,127.0.0.1,::1")
os.environ.setdefault("NO_PROXY", "localhost,127.0.0.1,::1")

# ===================== 日志 =====================
LOG_FILE = "contract_extractor.log"
logging.basicConfig(
    level=logging.DEBUG, format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
)

def global_excepthook(exc_type, exc_value, exc_traceback):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    logging.critical("程序崩溃", exc_info=(exc_type, exc_value, exc_traceback))
    with open("崩溃日志.txt", "w", encoding="utf-8") as f:
        f.write(f"时间: {datetime.now()}\n类型: {exc_type.__name__}\n信息: {exc_value}\n\n")
        traceback.print_exception(exc_type, exc_value, exc_traceback, file=f)
        f.write("\n请将此文件发给开发者\n")
sys.excepthook = global_excepthook

# ===================== 第三方库导入 =====================
import gradio as gr
import requests
from PIL import Image
import openpyxl

# ===================== Gradio schema 补丁 =====================
import gradio_client.utils as _gc_utils, functools as _ft
if not hasattr(_gc_utils, '_patched'):
    _orig = _gc_utils._json_schema_to_python_type
    @_ft.wraps(_orig)
    def _patched(schema, defs):
        return 'str' if isinstance(schema, bool) else _orig(schema, defs)
    _gc_utils._json_schema_to_python_type = _patched
    _gc_utils._patched = True

# ===================== Surya OCR 懒加载 =====================
SURYA_AVAILABLE = False
try:
    from surya import ocr as _surya_ocr
    from surya.model.recognition import RecognitionModel as _RecModel, RecognitionProcessor as _RecProc
    from surya.model.detection import DetectionModel as _DetModel, DetectionProcessor as _DetProc
    SURYA_AVAILABLE = True
    logging.info("Surya OCR 已就绪")
except Exception as _e:
    logging.warning(f"Surya OCR 不可用，降级使用视觉模型 OCR: {_e}")

_surya_instances = None
def get_surya():
    global _surya_instances
    if _surya_instances is None and SURYA_AVAILABLE:
        try:
            _surya_instances = (_DetModel(), _DetProc(), _RecModel(), _RecProc())
            logging.info("Surya OCR 模型加载完成")
        except Exception as e:
            logging.error(f"Surya 模型加载失败: {e}")
            return None
    return _surya_instances

# ===================== 配置 =====================
OLLAMA_BASE = "http://localhost:11434"
VISION_MODEL = "qwen2.5vl:3b"       # 视觉模型（OCR 降级用）
TEXT_MODEL = "qwen2.5:3b"           # 文本提取模型
EXCEL_PATH = "合同台账模板.xlsx"
MAX_IMAGE_LONGEST = 900
MAX_PDF_PAGES = 3
MAX_RETRIES = 3
RETRY_DELAY_SEC = 3

HEADERS = ["合同标题","甲方全称","乙方全称","合同总金额","金额大写",
           "生效日期","到期日期","服务期","付款条件","违约责任",
           "签约日期","合同标的","处理时间","状态"]

EXTRACT_PROMPT = """以下是合同内容，请从中提取结构化字段：

{text}

严格按照JSON格式返回（只返回JSON对象，不要其他文字）：

{
  "合同标题": "",
  "甲方全称": "",
  "乙方全称": "",
  "合同总金额": "",
  "金额大写": "",
  "生效日期": "",
  "到期日期": "",
  "服务期": "",
  "付款条件": "",
  "违约责任": "",
  "签约日期": "",
  "合同标的": ""
}

要求：
1. 每项填入找到的具体内容，找不到填 null
2. 金额带数字和单位
3. 日期统一 YYYY-MM-DD 格式
4. 只输出JSON，不要解释"""

# ===================== 工具函数 =====================

def check_ollama():
    try:
        r = requests.get(f"{OLLAMA_BASE}/api/tags", timeout=5)
        if r.status_code != 200:
            return False, f"Ollama 返回异常: HTTP {r.status_code}"
        models = r.json().get("models", [])
        names = [m["name"] for m in models]
        # 检查两个模型
        def fuzzy(name, installed):
            t = name.replace("-","").replace(":","").replace("latest","").lower()
            return any(t in n.replace("-","").replace(":","").replace("latest","").lower() for n in installed)
        vision_ok = fuzzy(VISION_MODEL, names)
        text_ok = fuzzy(TEXT_MODEL, names)
        parts = []
        if vision_ok: parts.append(f"视觉{VISION_MODEL}✅")
        else: parts.append(f"视觉{VISION_MODEL}❌")
        if text_ok: parts.append(f"文本{TEXT_MODEL}✅")
        else: parts.append(f"文本{TEXT_MODEL}❌")
        if not SURYA_AVAILABLE:
            if vision_ok:
                return True, f"{' '.join(parts)} (OCR:视觉模型降级)"
            return False, f"{' '.join(parts)} | 请运行: ollama pull {VISION_MODEL}"
        return True, f"Surya✅ {' '.join(parts)}"
    except requests.exceptions.ConnectionError:
        return False, "❌ Ollama 未启动"

def resize_image(img):
    if max(img.size) > MAX_IMAGE_LONGEST:
        ratio = MAX_IMAGE_LONGEST / max(img.size)
        img = img.resize((int(img.width*ratio), int(img.height*ratio)), Image.LANCZOS)
    return img

def pdf_to_images(pdf_path):
    import fitz
    doc = fitz.open(pdf_path)
    pages = min(len(doc), MAX_PDF_PAGES)
    images = []
    for i in range(pages):
        pix = doc[i].get_pixmap(matrix=fitz.Matrix(2, 2))
        images.append(Image.frombytes("RGB", [pix.width, pix.height], pix.samples))
    doc.close()
    return [resize_image(img) for img in images]

def img_to_b64(img):
    buf = BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return base64.b64encode(buf.getvalue()).decode("utf-8")

def safe_parse_json(text):
    if not text or not text.strip(): return None
    text = text.strip()
    try: return json.loads(text)
    except: pass
    cleaned = re.sub(r'^```(?:json)?\s*', '', text)
    cleaned = re.sub(r'\s*```$', '', cleaned).strip()
    try: return json.loads(cleaned)
    except: pass
    m = re.search(r'\{[\s\S]*\}', cleaned)
    if m:
        try: return json.loads(m.group())
        except: pass
    return None

def merge_results(results):
    if not results: return None
    if len(results) == 1: return results[0]
    merged = results[0].copy()
    for r in results[1:]:
        for key in merged:
            val = merged.get(key)
            if not val or str(val).strip() in ("", "null"):
                merged[key] = r.get(key, "")
    return merged

# ===================== ============= OCR 引擎 =====================

def ocr_via_vision(img, page_label=""):
    """降级方案：视觉模型 OCR"""
    b64 = img_to_b64(img)
    payload = {
        "model": VISION_MODEL,
        "messages": [{"role":"user", "content":"请从这张图中提取所有可见的文字内容，按原文顺序逐行输出，不要改写不要总结。", "images":[b64]}],
        "stream": False,
        "options": {"temperature": 0.0, "num_predict": 2048, "num_ctx": 2048},
    }
    for attempt in range(1, MAX_RETRIES+1):
        try:
            resp = requests.post(f"{OLLAMA_BASE}/api/chat", json=payload, timeout=180)
            if resp.status_code == 200:
                return resp.json().get("message",{}).get("content","")
        except:
            if attempt < MAX_RETRIES: time.sleep(RETRY_DELAY_SEC)
    return ""

def ocr_via_surya(img, page_label=""):
    """Surya OCR"""
    models = get_surya()
    if models is None: return None
    det_model, det_processor, rec_model, rec_processor = models
    try:
        predictions = _surya_ocr([img], [det_model], [det_processor], [rec_model], [rec_processor])
        lines = []
        for p in predictions:
            for line in p.text_lines:
                lines.append(line.text)
        return "\n".join(lines)
    except Exception as e:
        logging.error(f"Surya OCR 失败: {e}")
        return None

def ocr_image(img, page_label=""):
    """OCR 入口：优先 Surya，失败降级到视觉模型"""
    if SURYA_AVAILABLE:
        result = ocr_via_surya(img, page_label)
        if result: return result
        logging.warning(f"Surya 失败, 降级到视觉模型 OCR")
    return ocr_via_vision(img, page_label)

# ===================== 文本模型提取 =====================

def _call_ollama(payload):
    for attempt in range(1, MAX_RETRIES+1):
        try:
            resp = requests.post(f"{OLLAMA_BASE}/api/chat", json=payload, timeout=180)
            if resp.status_code == 200:
                return resp.json().get("message",{}).get("content","")
            logging.warning(f"Ollama 返回 {resp.status_code}: {resp.text[:200]}")
        except:
            if attempt < MAX_RETRIES: time.sleep(RETRY_DELAY_SEC)
    return ""

def extract_fields(text):
    payload = {
        "model": TEXT_MODEL,
        "messages": [{"role":"user", "content": EXTRACT_PROMPT.format(text=text)}],
        "stream": False, "format": "json",
        "options": {"temperature": 0.1, "num_predict": 1024},
    }
    return safe_parse_json(_call_ollama(payload))

# ===================== Excel =====================

def ensure_excel(path):
    p = Path(path)
    if not p.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = openpyxl.styles.Font(bold=True)
            c.alignment = openpyxl.styles.Alignment(horizontal="center")
        wb.save(path)

def append_row(data):
    ensure_excel(EXCEL_PATH)
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        row = ws.max_row + 1
        for i, key in enumerate(HEADERS[:-2], 1):
            val = data.get(key, "")
            if isinstance(val, (list, dict)): val = json.dumps(val, ensure_ascii=False)
            ws.cell(row=row, column=i, value=str(val) if val is not None else "")
        ws.cell(row=row, column=len(HEADERS)-1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row, column=len(HEADERS), value="成功")
        wb.save(EXCEL_PATH)
        return True, f"第{row}行"
    except PermissionError: return False, "Excel 被占用"
    except Exception as e: return False, f"写入失败: {e}"

# ===================== 核心处理 =====================

def process_single_file(file_path):
    """处理单个文件 -> {ocr_text, result_dict, log}"""
    path = Path(file_path)
    ext = path.suffix.lower()
    name = path.name
    try:
        if ext == ".pdf":
            images = pdf_to_images(file_path)
            if not images: return {"ocr":"", "result":None, "log":f"{name}: PDF 无内容"}
        elif ext in (".png",".jpg",".jpeg",".bmp",".tiff",".webp"):
            images = [resize_image(Image.open(file_path))]
        else: return {"ocr":"", "result":None, "log":f"{name}: 不支持格式"}

        # ---- OCR ----
        all_text = ""
        ocr_logs = []
        for i, img in enumerate(images):
            text = ocr_image(img, f"第{i+1}页")
            if text:
                all_text += f"\n=== 第{i+1}页 ===\n" + text
                ocr_logs.append(f"第{i+1}页✅")
            else:
                ocr_logs.append(f"第{i+1}页❌")

        if not all_text.strip():
            return {"ocr":"", "result":None, "log":f"{name}: OCR 失败 [{' '.join(ocr_logs)}]"}

        # ---- 提取 ----
        result = extract_fields(all_text)
        if not result:
            return {"ocr": all_text, "result":None, "log":f"{name}: 字段提取失败"}

        return {"ocr": all_text, "result": result, "log":f"{name}: OCR={' '.join(ocr_logs)} | 提取✅"}

    except Exception as e:
        return {"ocr":"", "result":None, "log":f"{name}: ❌ {e}"}

# ===================== Gradio UI =====================

CSS = """footer{display:none !important}"""

def build_ui():
    with gr.Blocks(title="合同智能提取工具", css=CSS, theme=gr.themes.Soft()) as demo:
        gr.Markdown("# 📄 合同智能提取工具\n上传合同 PDF/图片 → Surya OCR 提取原文 → AI 提取结构化字段 → 写入 Excel 台账")

        status = gr.Textbox(label="系统状态", interactive=False)

        # ---- 上传区 ----
        with gr.Row():
            file_input = gr.File(label="📎 上传合同文件（PDF / 图片）", file_count="multiple",
                                 file_types=[".pdf",".png",".jpg",".jpeg"], type="filepath", height=140)
            extract_btn = gr.Button("⚡ 开始处理", variant="primary", size="lg", scale=1)

        # ---- 文件列表 + 翻页 ----
        with gr.Row():
            file_checkbox = gr.CheckboxGroup(label="📋 合同列表（勾选后写入台账）", interactive=True, scale=4)
            prev_btn = gr.Button("◀ 上一份", size="sm", scale=1)
            next_btn = gr.Button("下一份 ▶", size="sm", scale=1)
            page_info = gr.Textbox(label="", value="0 / 0", interactive=False, scale=1)

        # ---- 左右分栏 ----
        with gr.Row(equal_height=False):
            with gr.Column(scale=1, min_width=400):
                ocr_box = gr.Textbox(label="📝 原文（OCR 结果）", lines=22, max_lines=40, interactive=False)
            with gr.Column(scale=1, min_width=400):
                result_table = gr.Dataframe(
                    headers=HEADERS[:-2],
                    label="📋 提取结果（双击单元格可手动修正）",
                    interactive=True, wrap=True,
                )

        # ---- 操作按钮 ----
        with gr.Row():
            save_all_btn = gr.Button("📥 写入全部勾选合同到台账", variant="primary", size="lg")
            save_current_btn = gr.Button("写入当前这份", size="lg")

        log_box = gr.Textbox(label="📝 处理日志", lines=6, max_lines=12)

        # ==================== 状态变量 ====================
        file_paths = gr.State([])                 # 所有文件路径
        current_idx = gr.State(0)                 # 当前显示第几个索引
        ocr_texts = gr.State({})                  # path → OCR 原文
        extract_results = gr.State({})            # path → 提取字段 dict
        extract_tables = gr.State({})             # path → 表格数据（二维列表）

        # ==================== 事件绑定 ====================
        demo.load(fn=lambda: check_ollama()[1], outputs=[status])

        # ---- 上传后更新文件列表 ----
        def on_upload(files):
            if not files: return [], [], 0, "0 / 0", "", [], {}
            paths = [f.name for f in files]
            choices = [Path(f.name).name for f in files]
            # 默认全选
            return paths, choices, 0, f"1 / {len(choices)}", "", [], {}
        file_input.change(fn=on_upload, inputs=[file_input],
                          outputs=[file_paths, file_checkbox, current_idx, page_info, ocr_box, result_table, extract_results])

        # ---- 开始处理 ----
        def on_extract(files, paths_state, idx_state, ocr_state, extract_state):
            if not paths_state: return "", [], "请先上传文件", "", 0, "0 / 0", ocr_state, extract_state

            new_ocr = dict(ocr_state) if ocr_state else {}
            new_extract = dict(extract_state) if extract_state else {}
            new_tables = {}
            logs = []

            total = len(paths_state)
            for i, p in enumerate(paths_state):
                name = Path(p).name
                logs.append(f"[{i+1}/{total}] {name} 处理中...")
                if p in new_ocr and p in new_extract:
                    logs[-1] = f"[{i+1}/{total}] {name} 已处理，跳过"
                    continue
                res = process_single_file(p)
                new_ocr[p] = res["ocr"]
                if res["result"]:
                    new_extract[p] = res["result"]
                    tbl_row = [res["result"].get(h, "") for h in HEADERS[:-2]]
                    new_tables[p] = [tbl_row]
                logs[-1] = f"[{i+1}/{total}] {res['log']}"

            # 显示第一个
            first_path = paths_state[0]
            first_ocr = new_ocr.get(first_path, "")
            first_tbl = new_tables.get(first_path, [])
            page_txt = f"1 / {total}" if total > 0 else "0 / 0"

            return first_ocr, first_tbl, "\n".join(logs), paths_state, 0, page_txt, new_ocr, new_extract

        extract_btn.click(fn=on_extract,
            inputs=[file_input, file_paths, current_idx, ocr_texts, extract_results],
            outputs=[ocr_box, result_table, log_box, file_paths, current_idx, page_info, ocr_texts, extract_results])

        # ---- 翻页 ----
        def go_prev(idx_state, paths_state, ocr_state, extract_state):
            if not paths_state or idx_state <= 0: return "", [], "0 / 0", idx_state
            new_idx = idx_state - 1
            p = paths_state[new_idx]
            ocr_t = (ocr_state or {}).get(p, "")
            tbl = _get_table(extract_state, p)
            return ocr_t, tbl, f"{new_idx+1} / {len(paths_state)}", new_idx

        def go_next(idx_state, paths_state, ocr_state, extract_state):
            if not paths_state or idx_state >= len(paths_state)-1: return "", [], "0 / 0", idx_state
            new_idx = idx_state + 1
            p = paths_state[new_idx]
            ocr_t = (ocr_state or {}).get(p, "")
            tbl = _get_table(extract_state, p)
            return ocr_t, tbl, f"{new_idx+1} / {len(paths_state)}", new_idx

        def _get_table(extract_state, path):
            if not extract_state or path not in extract_state: return []
            r = extract_state[path]
            return [[r.get(h, "") for h in HEADERS[:-2]]]

        prev_btn.click(fn=go_prev, inputs=[current_idx, file_paths, ocr_texts, extract_results],
                       outputs=[ocr_box, result_table, page_info, current_idx])
        next_btn.click(fn=go_next, inputs=[current_idx, file_paths, ocr_texts, extract_results],
                       outputs=[ocr_box, result_table, page_info, current_idx])

        # ---- 写入台账 ----
        def on_save_all(current_tbl, paths_state, checkbox_state, idx_state, extract_state):
            """写入全部勾选合同"""
            if not checkbox_state: return "请勾选要写入的合同"
            # checkbox_state 是文件名列表
            selected_names = set(checkbox_state)
            logs = []
            for p in (paths_state or []):
                name = Path(p).name
                if name not in selected_names:
                    continue
                if extract_state and p in extract_state:
                    ok, msg = append_row(extract_state[p])
                    logs.append(f"{name}: {msg}")
                else:
                    logs.append(f"{name}: 无提取结果")
            return "\n".join(logs) if logs else "没有选中任何已处理的合同"

        def on_save_current(current_tbl, idx_state, paths_state, extract_state):
            """写入当前这份"""
            if not paths_state or idx_state >= len(paths_state):
                return "没有当前合同"
            p = paths_state[idx_state]
            if not extract_state or p not in extract_state:
                return "当前合同无提取结果"
            ok, msg = append_row(extract_state[p])
            return f"{Path(p).name}: {msg}"

        save_all_btn.click(fn=on_save_all,
            inputs=[result_table, file_paths, file_checkbox, current_idx, extract_results],
            outputs=[log_box])
        save_current_btn.click(fn=on_save_current,
            inputs=[result_table, current_idx, file_paths, extract_results],
            outputs=[log_box])

        gr.Markdown("---\n**使用提示** ①上传文件 → ②点击开始处理 → ③左右翻页查看 → ④修正字段 → ⑤勾选后写入台账")

    return demo

# ===================== 入口 =====================
if __name__ == "__main__":
    try:
        ensure_excel(EXCEL_PATH)
        ok, msg = check_ollama()
        logging.info(f"状态: {msg}")
        print(f"\n{'='*50}\n  合同智能提取工具\n  状态: {msg}\n  打开: http://127.0.0.1:7860\n  日志: {LOG_FILE}\n{'='*50}\n")
        demo = build_ui()
        demo.launch(server_name="127.0.0.1", server_port=7860, inbrowser=True, share=False, quiet=False, show_error=True)
    except Exception as e:
        logging.exception("启动失败")
        with open("崩溃日志.txt", "w", encoding="utf-8") as f:
            f.write(f"时间: {datetime.now()}\n错误: {e}\n\n"); traceback.print_exc(file=f)
        print(f"\n❌ 启动失败，查看崩溃日志.txt")
        try: input("按 Enter 退出...")
        except: pass
